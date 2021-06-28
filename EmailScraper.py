"""
Email Scraper

This script used to scrape the text from all the emails
in a Gmail account that is subscribed to various humanitarian
newsletters and charities.  The text is then parsed for the frequency of certain
search terms and then recorded in the accompanying Excel Workbook.  This tool
was used to analyze a large volume of emails to paint a picture of how certain
organizations use words that connote positive or negative emotions and direct
solicitations in their fundraising outreach.

"""

import openpyxl
import imapclient
import imaplib
import pyzmail
import re
import os

#  Increase memory allowance
imaplib._MAXLINE = 10000000

# Load Excel Workbook
wb = openpyxl.load_workbook(os.getcwd()+"\\WordFrequenciesData.xlsx")
sheet = wb.get_sheet_by_name('Sheet1')


# Regex to search each email by the search word in the cell and output to a frequency cell
def excelRE(Email, searchCell, freqCell):
    searchTerm = r'\b' + searchCell.value.lower() + r'\b'
    freqCell.value = len(re.compile(searchTerm).findall(Email))


# Log into Gmail
myemail = input('Please enter your email address: ')
password = input('Please enter your password: ')
imapObj = imapclient.IMAPClient('imap.gmail.com', ssl=True)
imapObj.login(myemail, password)

# Returns unique IDs for each email
imapObj.select_folder('INBOX', readonly=True)
allEmailUID = imapObj.search(['ALL'])

# Loop through all emails
for email in allEmailUID:
    rawMessage = imapObj.fetch(email, ['BODY[]'])
    # Create PyzMessage Object
    try:
        message = pyzmail.PyzMessage.factory(rawMessage[email][b'BODY[]'])
    except KeyError:
        message = pyzmail.PyzMessage.factory(rawMessage[email]['BODY[]'])
    # Parse Message text
    try:
    # Get message text if email is plaintext
        messageText = message.text_part.get_payload().decode(message.text_part.charset).lower()
    except AttributeError:
    # Get message text if email is HTML
        messageText = message.html_part.get_payload().decode(message.html_part.charset).lower()
    # Add Email address,Sender, Subject to Excel workbook
    sheet['A' + str(allEmailUID.index(email)+3)].value = message.get_addresses('from')[0][1]
    sheet['B' + str(allEmailUID.index(email)+3)].value = message.get_addresses('from')[0][0]
    sheet['C' + str(allEmailUID.index(email)+3)].value = message.get_subject()
    for column in range(4, 25):
    # Call Regex function to populate frequency of search words in each column
    # Row 2 contains the search terms separated by column
        excelRE(messageText, sheet.cell(row=2, column=column), sheet.cell(row=allEmailUID.index(email)+3, column=column))

# Save workbook
wb.save(os.getcwd()+"\\WordFrequenciesData.xlsx")
wb.close()
